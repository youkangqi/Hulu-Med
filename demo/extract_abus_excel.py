import argparse
import csv
from pathlib import Path

'''
遍历 outputs/abus_reports_raw_report_cn 下每个病人目录，提取：

labelled_infer.txt（多张图：保存 image_index + 输出文本）
unlabelled_infer.txt（多张图：保存 image_index + 输出文本）
multi_images_report.txt
multi_images_report_labels_only.txt

输出为 CSV（默认）：
outputs_extracted.csv

运行示例
python demo/abus_outputs_extract.py --input-dir outputs/abus_reports_raw_report_cn_with_keys --output-csv outputs/abus_reports_raw_report_cn_with_keys/outputs_extracted.csv
'''

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract non-empty B column values and corresponding G column values."
    )
    parser.add_argument(
        "--input-path",
        default="data/abus_data/abus.xlsx",
        help="Path to the Excel file.",
    )
    parser.add_argument(
        "--output-path",
        default="outputs/abus_b_g.csv",
        help="Path to write the extracted CSV.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name (defaults to active sheet).",
    )
    return parser.parse_args()


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def main():
    args = parse_args()
    input_path = Path(args.input_path)
    output_path = Path(args.output_path)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to read .xlsx files. Install with: pip install openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(input_path, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["row", "col_B", "col_G"])
        for row_idx in range(1, sheet.max_row + 1):
            b_value = normalize_cell(sheet.cell(row=row_idx, column=2).value)
            if not b_value:
                continue
            g_value = normalize_cell(sheet.cell(row=row_idx, column=7).value)
            writer.writerow([row_idx, b_value, g_value])

    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
